# -*- encoding: utf-8 -*-
import pandas
from openpyxl.workbook import Workbook
from openpyxl import load_workbook
data = pandas.read_csv("exceldata/testing_publix_salesv2.csv")

wb = load_workbook('actual_sales_list.xlsx')
ws = wb.active

# if ws['D3'] == "A1FE":
column_a = ws['A']
column_b = ws['B']
column_c = ws['C']
for cell, cell2, cell3 in zip(column_a, column_b, column_c):
    if cell.value is not None and cell2.value is not None and cell3.value is not None:
        print(f"THIS IS ON AISLE {cell2.value}\n"
              f"<li class='product-item'>\n"
              f"   <span class='product-name'>{cell.value}</span>\n"
              f"   <span class='secondary-info'><b>ALSO ON:</b> {cell3.value}</span>\n"
              f"</li>")
# print(data["Item"])
# print(data["Display Name"])

